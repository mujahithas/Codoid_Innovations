import openpyxl


class Openpyxl:

    def get_count_row(self, file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        return sheet.max_row

    def get_count_column(self, file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        return (sheet.max_column)

    def read_data(self, file, sheet_name, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        return sheet.cell(row=rownum, column=columnno).value

